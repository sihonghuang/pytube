# 获取以下内容并写入excel
# 作者/标题/浏览量/日期/描述/链接
import re
import sys
import time
from urllib.error import HTTPError, URLError
import openpyxl as xl
import requests
import self as self
import youtubeUtil
from pytube import Channel
from pytube import YouTube
from pytube.cli import on_progress
from sihong.Utils import commonUtil


def get_path():
    if sys.platform.startswith('win'):
        return 'D:/'  # Windows系统路径
    elif sys.platform.startswith('darwin') or sys.platform.startswith('linux'):
        return '/Users/sihonghuang/Documents/'  # macOS和Linux系统路径
    else:
        return ''  # 其他操作系统路径（可以根据需求进行适当修改）


default_ptah = get_path() + "油管资料/"  # excel默认保存露营 # TODO


def get_yt(url):
    while True:
        try:
            yt = YouTube(url, on_progress_callback=on_progress)
            return yt
            break
        except HTTPError:
            self.logger.error("请求出错一次：HTTPError")
            continue
        except URLError:
            self.logger.error("请求出错一次：URLError")
            continue


def downloadYoutube(url):
    yt = get_yt(url)
    fuchsia = '\033[38;2;255;00;255m'  # color as hex #FF00FF
    reset_color = '\033[39m'
    # 配置信息
    author = commonUtil.remove_symbol(yt.author)  # 视频作者（不带空格）
    title = commonUtil.remove_symbol(yt.title)  # 视频标题(不带空格)
    print("视频上传时间: " + yt.publish_date.strftime(
        "%Y%m%d") + "\n" + "今天的日期: " + time.strftime("%Y%m%d"))
    save_path = save_path_check(channel_type(author))
    description = youtubeUtil.get_video_description(url)
    file_path = author + "/" + yt.publish_date.strftime("%Y-%m-%d") + "_" + title
    print("文件保存路径:" + save_path + file_path)
    print(f'\n' + fuchsia + '开始下载: ', author + "----" + title, '~ viewed', yt.views,
          'times.')
    # 下载前先检查是否有该路径，没有则创建
    commonUtil.file_path_check(save_path, file_path)
    start = time.time()
    # 下载字幕
    subtitle = youtubeUtil.download_caption(url, save_path, file_path, title)
    # 下载封面
    video_img = youtubeUtil.request_download(yt.thumbnail_url, save_path, file_path,
                                             title)  # 下载视频封面

    # 下载视频
    #video_num = youtubeUtil.download_video(yt, save_path, file_path)
    video_num = youtubeUtil.download_highest_resolution_video(url,save_path,file_path)
    # 下载音频 暂时用youtube-dl的下载方法
    audio = False#youtubeUtil.download_audio(yt, save_path, file_path)
    # 写入excel #转移到下载视频里
    end = time.time()
    if video_img is True or video_num != 0 or audio is True:
        write_to_excel(yt.author, yt.title, yt.views, yt.publish_date, description, url,author)
        print(f'\n完成下载:  {title}' + "下载完成,共计花费了{:.2f}秒".format(end - start) + reset_color)
    else:
        print(f'\n====================本集视频素材均已下载====================')
    return video_num


def author_name_check(author):
    if author == "Forest Film: camping, off grid cabin and bushcraft":
        author = "Forest Film"
        return author
    else:
        return author


def download_multi_video(urls):
    num = 1
    for url in urls:
        print("开始下载第" + str(num) + "个视频")
        downloadYoutube(url)
        num = num + 1
    print("批量下载完成，总共下载" + str(num) + "个视频")


# 批量下载
def download_by_channels(channe_lists):
    url = 'https://www.youtube.com/@'
    url_end = '/videos'
    num = 0
    sum_video_num = 0  # 本次总下载数
    up = set()
    download_num = 6  # 循环下载up主的视频数
    for channel in channe_lists:
        video_num = 0  # 初始化视频下载数
        num += 1  # 初始化下载的推主序号
        download_url = url + channel + url_end
        print(download_url)
        yt_channel = Channel(download_url)
        print(yt_channel.video_urls[:download_num])
        print("开始下载第" + str(num) + "个推主:" + yt_channel.channel_name + " 最新的" + str(
            download_num) + "个视频")
        if len(yt_channel.video_urls) != 0:
            for download_url in yt_channel.video_urls[:download_num]:
                try:
                    video_num = video_num + downloadYoutube(download_url)
                    if video_num != 0:
                        up.add(yt_channel.channel_name)  # Use add() to add unique values to the set
                    time.sleep(3)
                except Exception as e:
                    print(e)
                finally:
                    continue
        sum_video_num += video_num
        time.sleep(5)  # 每个推主中间停顿下
    print("批量下载完成，共查找" + str(
        num) + "个UP主,其中:" + ", ".join(up) + "有新视频，总共下载下载了 " + str(
        sum_video_num) + " 视频。")


# 根据不同类型的推主放到不同文件夹
def save_path_check(channel_type):
    commonUtil.file_path_check(default_ptah, channel_type)
    result = default_ptah + channel_type + '/'
    return result


# 根据不同类型的推主放到不同excel
def excel_path_check(channel_type):
    excel_path = default_ptah + channel_type + '.xlsx'
    return excel_path


# 视频风格
def channel_type(channel_name):
    category_lists = {
        '露营': ['HikeCampClimb', 'AtikAilesi', 'BaumOutdoors', 'BahadirKlc',
                 'tahtarizkyorigma', 'WoodsOfSilence',
                 'Lonewolfwildcamping', 'ForestFilm', 'XanderBudnick',
                 'SBWildernessAdventures', 'BrooksandBirches',
                 'tabi-ie', 'HighlandWoodsman', 'Kampkolik', 'joinmeoutdoors',
                 '2withnature', 'forestsolitude7448',
                 'silentfamily', 'jaylegere', '365GnDoadayz', 'LeavesDiary88'],
        '徒步': ['HarmenHoek', 'kraigadams'],
        '户外': ['ThisIsMyAlaska', 'BushcraftAdventure', 'NorwegianXplorer', 'Survivaland',
                 'MyMethead'],
        '街景': ['IntotheWildFilms'],
        '相机': ['snapsbyfox', 'MannyOrtiz', 'benjhaisch', 'LeeZavitz']
    }
    for category, channel_list in category_lists.items():
        if channel_name in channel_list:
            return category
    return None  # 如果没有匹配项，返回None或其他适当的值


def today_release_youtuber_list(url, download_type, channel_name):
    yt = get_yt(url)
    fuchsia = '\033[38;2;255;00;255m'  # color as hex #FF00FF
    reset_color = '\033[39m'
    author = commonUtil.remove_symbol(yt.author)  # 视频作者（不带空格）
    title = commonUtil.remove_symbol(yt.title)  # 视频标题(不带空格)


def write_to_excel(author, title, views, publish_date, description, url, channel_name):
    try:
        excel_save_path = excel_path_check(channel_type(channel_name))
        wb = xl.load_workbook(excel_save_path)  # 获取文件
        ws = wb[channel_name]
        nrows = ws.max_row + 1  # 总行数
        ws.cell(row=nrows, column=1, value=author)
        ws.cell(row=nrows, column=2, value=commonUtil.googleTrans(title, "英文"))
        ws.cell(row=nrows, column=3, value=views)
        ws.cell(row=nrows, column=4, value=publish_date.strftime("%Y%m%d"))
        ws.cell(row=nrows, column=5, value=description + "\r\n" + commonUtil.googleTrans(description, "中文"))
        ws.cell(row=nrows, column=6, value=author + " | " + commonUtil.googleTrans(title, "中文") + "(第 期)")
        ws.cell(row=nrows, column=7, value=url + "\r\n转自【" + author + "】" + "\r\n" + commonUtil.googleTrans(description, "中文"))
        wb.save(excel_save_path)
        print("写入表格行 " + str(nrows) + ": " + author + title)
    except Exception as e:
        # 在这里处理异常，可以打印异常信息，也可以进行其他操作
        print("发生异常：" + str(e))


def get_user_choice(channel_list):
    while True:
        print("选择要使用的功能：\r\n" + "0. 下载单个视频" + "\r\n" + "1. 批量下载视频")
        choice = input("输入选项编号：")  # 0 单个视频 1 批量
        if choice.isdigit():
            choice = int(choice)
            if choice == 0:
                input_url = input("输入视频地址：")
                downloadYoutube(input_url)
                print("下载链接:" + input_url)
                break

            # 查找每个频道最新的3个视频
            elif choice == 1:
                download_by_channels(channel_list)
                break
        print("无效的选项，请重新输入。")


if __name__ == '__main__':
    channel_list = ['AtikAilesi','365GunDogadayiz']
    get_user_choice(channel_list)

    # try:
    #     # 下载单个视频  
    #     input_url=sys.argv[1]#视频连接
    #     downloadYoutube(input_url, 0, '')
    # except Exception as e:
    #     print(e)
